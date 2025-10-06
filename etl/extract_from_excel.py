from __future__ import annotations
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter

def _excel_used_range(path: str, sheet_name: str, start_cell: str):
    wb = load_workbook(path, data_only=True, read_only=True, keep_vba=True)
    ws = wb[sheet_name]

    col_letters = ''.join(ch for ch in start_cell if ch.isalpha()).upper()
    row_digits = ''.join(ch for ch in start_cell if ch.isdigit())
    start_row = int(row_digits)
    start_col = column_index_from_string(col_letters)

    last_row = start_row
    last_col = start_col 

    for r in ws.iter_rows(min_row=start_row, min_col=start_col, values_only=True):
        if any(v not in (None, "") for v in r):
            last_row += 1
            for idx, v in enumerate(r, start=start_col):
                if v not in (None, "") and idx > last_col:
                    last_col = idx
    
    last_row = max(last_row - 1, start_row)

    return start_row, start_col, last_row, last_col

def _normalize_col_name(s: str) -> str:
    return(
        str(s)
        .replace("\n", " ")
        .strip()
        .lower()
        .replace("  ", " ")
        .replace(" ", "_")
        .replace("-", "_")
        .replace("[", "")
        .replace("]", "")
    )

def read_table_from_excel(path: str, sheet_name: str, start_cell: str="D12", normalize_headers: bool = True) -> pd.DataFrame:
    sr, sc, er, ec = _excel_used_range(path, sheet_name, start_cell)

    start_col_letter = get_column_letter(sc)
    end_col_letter = get_column_letter(ec)
    usecols = f"{start_col_letter}:{end_col_letter}"

    header_row_index = sr - 1

    nrows = er - (sr - 1)
    if nrows < 1:
        raise ValueError(
            f"Computed nrows < 1 (sr={sr}, er={er}). Check sheet '{sheet_name}' and start_cell '{start_cell}'."
        )
    
    df = pd.read_excel(
        path,
        sheet_name=sheet_name,
        engine="openpyxl",
        header=header_row_index,
        usecols=usecols,
        nrows=nrows
    )

    df = df.dropna(how="all")
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df.columns = [str(c).strip() for c in df.columns]
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [" ".join(str(x) for x in tup if str(x) != "nan").strip() for tup in df.columns]

    if normalize_headers:
        df.columns = [_normalize_col_name(c) for c in df.columns]

    def _dedupe(names):
        seen = {}
        out = []
        for name in names:
            if name not in seen:
                seen[name] = 0
                out.append(name)
            else:
                seen[name] += 1
                out.append(f"{name}_{seen[name]}")
        
        return out

    orig_cols = list(df.columns)
    df.columns = _dedupe(df.columns)

    assert df.columns.is_unique, "Column names are still not unique after de-duplication."
    return df

